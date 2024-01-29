import openpyxl
# file=(r"C:\Users\veenu\Downloads\Data_Driven.xlsx")
# wb=openpyxl.load_workbook(file)
# Sheet=wb["Sheet1"]
# nth_row=Sheet.max_row
# nth_column=Sheet.max_row
# print(Sheet.max_row)
# print(Sheet.max_column)
# Sheet.cell(5,1).value=4
# Sheet.cell(5,2).value="Radha"
# Sheet.cell(5,3).value=400
# print(Sheet.cell(2,2).value)

# for r in range(1,nth_row+1):
#     for c in range(1, nth_column + 1):
#         print(Sheet.cell(r,c).value)
#     print("\n")

# wb.save(file)
file1=r"C:\Users\veenu\Downloads\sample.xlsx"
wb1=openpyxl.Workbook()
wb1.create_sheet("Sample")
print(wb1.active)
sheet=wb1["Sample"]
wb1.active=sheet
sheet.cell(1,1).value="data"
print(wb1.sheetnames)
sheet.title="Create Sheet"
#wb1.remove("Sheet")
wb1.save(file1)
# wb=openpyxl.Workbook()
# wb.create_sheet("sheet_created")
# print(wb.active)
# sheet = wb["sheet_created"]
# wb.active=sheet
# print(wb.active)
# sheet.cell(1,1).value="data"
# wb.save(file)


