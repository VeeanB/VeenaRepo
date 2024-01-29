import openpyxl
from openpyxl.styles import PatternFill
def getRow(xlpath,sheet):
    wb=openpyxl.load_workbook(r"C:\Users\veenu\PycharmProjects\PythonClass\selenium\Practice_PyTestFrame\Test Data\TestData_DDT.xlsx")
    sheet=wb["Sheet1"]
    return sheet.max_row
def getCol(xlpath,sheet):
    wb=openpyxl.load_workbook(r"C:\Users\veenu\PycharmProjects\PythonClass\selenium\Practice_PyTestFrame\Test Data\TestData_DDT.xlsx")
    sheet=wb["Sheet1"]
    return sheet.max_column
def getValue(xlpath,sheet,r_no,c_no):
    wb = openpyxl.load_workbook(r"C:\Users\veenu\PycharmProjects\PythonClass\selenium\Practice_PyTestFrame\Test Data\TestData_DDT.xlsx")
    sheet = wb["Sheet1"]
    return sheet.cell(r_no,c_no).value
def writeValue(xlpath,sheet,r_no,c_no,data):
    wb = openpyxl.load_workbook(r"C:\Users\veenu\PycharmProjects\PythonClass\selenium\Practice_PyTestFrame\Test Data\TestData_DDT.xlsx")
    sheet = wb["Sheet1"]
    sheet.cell(r_no,c_no).value=data
    wb.save(r"C:\Users\veenu\PycharmProjects\PythonClass\selenium\Practice_PyTestFrame\Test Data\TestData_DDT.xlsx")
def fillRed(xlpath,sheet,r_no,c_no):
    wb = openpyxl.load_workbook(r"C:\Users\veenu\PycharmProjects\PythonClass\selenium\Practice_PyTestFrame\Test Data\TestData_DDT.xlsx")
    sheet = wb["Sheet1"]
    red = PatternFill(start_color='eb0905', end_color='eb0905', fill_type='solid')
    sheet.cell(r_no,c_no).fill = red
    wb.save(r"C:\Users\veenu\PycharmProjects\PythonClass\selenium\Practice_PyTestFrame\Test Data\TestData_DDT.xlsx")
def fillGreen(xlpath,sheet,r_no,c_no):
    wb = openpyxl.load_workbook(r"C:\Users\veenu\Downloads\Simple_Interest.xlsx")
    sheet = wb["Sheet1"]
    green=PatternFill(start_color='60b212',end_color='60b212',fill_type='solid')
    sheet.cell(r_no,c_no).fill = green
    wb.save(r"C:\Users\veenu\PycharmProjects\PythonClass\selenium\Practice_PyTestFrame\Test Data\TestData_DDT.xlsx")


